from fastapi import FastAPI, HTTPException, Query, Body, Request, Form, File, UploadFile
from fastapi.responses import RedirectResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import List, Optional
from fastapi.staticfiles import StaticFiles
import os
import logging
import database
import csv
import io
import uuid
import shutil
import smtplib
import ssl
import base64
import pandas as pd
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from fastapi.responses import RedirectResponse, StreamingResponse
from dotenv import load_dotenv
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

load_dotenv()

GMAIL_SCOPES = ['https://www.googleapis.com/auth/gmail.send']

def get_gmail_service():
    """Returns an authorized Gmail API service instance."""
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', GMAIL_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
            with open('token.json', 'w') as token:
                token.write(creds.to_json())
        else:
            return None  # No token yet, must run setup_gmail.py
    return build('gmail', 'v1', credentials=creds)

logging.basicConfig(level=logging.INFO)

app = FastAPI(title="Cuentas Por Pagar API")

# Ensure static folder exists
os.makedirs("static", exist_ok=True)
app.mount("/static", StaticFiles(directory="static", html=True), name="static")

@app.get("/")
async def root():
    return RedirectResponse(url="/static/index.html")

class PlanPagoRequest(BaseModel):
    nros_unicos: List[int]
    fecha_planificada: str
    banco: str

class ForecastEventRequest(BaseModel):
    fecha: str
    tipo_evento: str
    valor: float = 1.0

# --- Modelos de Gastos Programados ---
class ExpenseTemplateRequest(BaseModel):
    id: Optional[int] = None
    descripcion: str
    tipo: str
    monto_estimado_usd: float
    dia_mes_estimado: int

class ProgrammedExpense(BaseModel):
    id: Optional[int] = None
    descripcion: str
    tipo: str
    monto_usd: float
    fecha_proyectada: str
    estado: str = "Pendiente"

class BatchExpenseRequest(BaseModel):
    mes: int
    anio: int
    gastos: List[ProgrammedExpense]

# --- Modelos Módulo Pagos e Indexación ---
class ProveedorCondicion(BaseModel):
    CodProv: str
    DiasNoIndexacion: Optional[int] = 0
    BaseDiasCredito: Optional[str] = "EMISION"
    DiasVencimiento: Optional[int] = 0
    ProntoPago1_Dias: Optional[int] = 0
    ProntoPago1_Pct: Optional[float] = 0.0
    ProntoPago2_Dias: Optional[int] = 0
    ProntoPago2_Pct: Optional[float] = 0.0
    Email: Optional[str] = None

class InvoiceReference(BaseModel):
    CodProv: str
    NumeroD: str
    MontoRetencionBs: Optional[float] = 0.0

class DebitNoteActionRequest(BaseModel):
    Invoices: List[InvoiceReference]

class DebitNoteRegisterRequest(BaseModel):
    Invoices: List[InvoiceReference]
    NotaDebitoID: str

class AbonoRequest(BaseModel):
    NumeroD: str
    CodProv: str
    FechaAbono: str
    MontoBsAbonado: float
    TasaCambioDiaAbono: float
    MontoUsdAbonado: float
    AplicaIndexacion: bool
    Referencia: Optional[str] = ""

@app.post("/api/plan-pagos")
async def planificar_pagos(payload: PlanPagoRequest):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Insert each record into the PagosPlanificados table
        for nro in payload.nros_unicos:
            # Check if it already exists, if so update it, otherwise insert
            check_query = "SELECT ID FROM EnterpriseAdmin_AMC.Procurement.PagosPlanificados WHERE NroUnico = ?"
            cursor.execute(check_query, (nro,))
            existing = cursor.fetchone()
            
            if existing:
                update_query = """
                    UPDATE EnterpriseAdmin_AMC.Procurement.PagosPlanificados
                    SET FechaPlanificada = ?, Banco = ?, CodUsua = 'API_USER'
                    WHERE NroUnico = ?
                """
                cursor.execute(update_query, (payload.fecha_planificada, payload.banco, nro))
            else:
                insert_query = """
                    INSERT INTO EnterpriseAdmin_AMC.Procurement.PagosPlanificados 
                    (NroUnico, FechaPlanificada, Banco, CodUsua)
                    VALUES (?, ?, ?, 'API_USER')
                """
                cursor.execute(insert_query, (nro, payload.fecha_planificada, payload.banco))
        
        conn.commit()
        return {"message": f"Successfully planned {len(payload.nros_unicos)} payment(s)."}
        
    except Exception as e:
        logging.error(f"Error saving planned payments: {e}", exc_info=True)
        if 'conn' in locals():
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

@app.post("/api/plan-pagos/unplan", response_model=None)
async def unplan_pagos(payload: dict = Body(...)):
    try:
        nros_unicos = payload.get("nros_unicos", [])
        if not nros_unicos:
            return {"message": "No invoices provided to unplan."}
            
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        placeholders = ','.join(['?'] * len(nros_unicos))
        delete_query = f"DELETE FROM EnterpriseAdmin_AMC.Procurement.PagosPlanificados WHERE NroUnico IN ({placeholders})"
        cursor.execute(delete_query, nros_unicos)
        
        conn.commit()
        return {"message": f"Successfully unplanned {cursor.rowcount} payment(s)."}
    except Exception as e:
        logging.error(f"Error unplanning payments: {e}", exc_info=True)
        if 'conn' in locals():
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/cuentas-por-pagar")
async def get_cuentas_por_pagar(search: str = Query("", description="Search term for NumeroD or Provider"), desde: Optional[str] = Query(None), hasta: Optional[str] = Query(None)):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        date_filter = ""
        date_params = []
        if desde:
            date_filter += " AND CAST(SAACXP.FechaE AS DATE) >= ?"
            date_params.append(desde)
        if hasta:
            date_filter += " AND CAST(SAACXP.FechaE AS DATE) <= ?"
            date_params.append(hasta)
            
        if not desde and not hasta:
            date_filter = " AND SAACXP.FechaE >= DATEADD(month, -4, GETDATE())"
            
        # Cleanup: Remove from planning any invoice that has been fully paid (Saldo near 0)
        cleanup_query = """
            DELETE PP
            FROM EnterpriseAdmin_AMC.Procurement.PagosPlanificados PP
            INNER JOIN EnterpriseAdmin_AMC.dbo.SAACXP ON PP.NroUnico = SAACXP.NroUnico
            WHERE SAACXP.Saldo <= 0.01 AND SAACXP.TipoCxP = '10'
        """
        cursor.execute(cleanup_query)
        conn.commit()
        
        # Build query
        query = f"""
            SELECT
              SACOMP.FechaI,
              SACOMP.FechaE,
              SACOMP.FechaV,
              SAPROV.Descrip,
              SAACXP.RetenIVA,
              SAACXP.SaldoAct,
              SAACXP.Monto,
              SAACXP.CodOper,
              SAACXP.MontoNeto,
              SAACXP.Saldo,
              SAACXP.MtoTax,
              SACOMP.MtoPagos,
              SACOMP.SaldoAct AS SaldoAct_SACOMP,
              SACOMP.MtoNCredito,
              SACOMP.MtoNDebito,
              SACOMP.Signo,
              SACOMP.NumeroD AS NumeroD_SACOMP,
              SAACXP.NroCtrol,
              SACOMP.MtoTotal,
              SACOMP.Contado,
              SACOMP.Credito,
              SAACXP.NroUnico,
              SAACXP.CodSucu,
              SAACXP.CodProv,
              SAACXP.NumeroD,
              SACOMP.CodSucu AS CodSucu_SACOMP,
              SACOMP.TipoCom,
              SACOMP.Notas10,
              SAPAGCXP.NumeroD AS NumeroD_SAPAGCXP,
              dt_emision.dolarbcv AS TasaEmision,
              dt_actual.dolarbcv AS TasaActual,
              PP.ID AS Plan_ID,
              PP.Banco AS Plan_Banco,
              PP.FechaPlanificada AS Plan_Fecha,
              CAST(CASE WHEN SAACXP.RetenIVA > 0 THEN 1 ELSE 0 END AS BIT) AS Has_Retencion,
              CAST(CASE WHEN abonos.TotalBs IS NOT NULL THEN 1 ELSE 0 END AS BIT) AS Has_Abonos,
              ISNULL(abonos.TotalBs, 0) AS TotalBsAbonado
            FROM dbo.SAACXP
            OUTER APPLY (
                SELECT SUM(MontoBsAbonado) AS TotalBs
                FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos A 
                WHERE A.CodProv = SAACXP.CodProv AND A.NumeroD = SAACXP.NumeroD
            ) abonos
            OUTER APPLY (
                SELECT TOP 1 NumeroD
                FROM dbo.SAPAGCXP
                WHERE SAPAGCXP.NroUnico = SAACXP.NroUnico
            ) SAPAGCXP
            LEFT OUTER JOIN dbo.SAPROV ON SAACXP.CodProv = SAPROV.CodProv
            LEFT OUTER JOIN dbo.SAIPACXP ON SAACXP.NroUnico = SAIPACXP.NroUnico
            LEFT OUTER JOIN dbo.SACOMP ON SAACXP.NumeroD = SACOMP.NumeroD AND SAACXP.CodProv = SACOMP.CodProv
            OUTER APPLY (
                SELECT TOP 1 dolarbcv 
                FROM EnterpriseAdmin_AMC.dbo.dolartoday 
                WHERE CAST(fecha AS DATE) <= CAST(SAACXP.FechaE AS DATE)
                ORDER BY fecha DESC
            ) dt_emision
            OUTER APPLY (
                SELECT TOP 1 dolarbcv 
                FROM EnterpriseAdmin_AMC.dbo.dolartoday 
                WHERE dolarbcv IS NOT NULL
                ORDER BY id DESC
            ) dt_actual
            LEFT OUTER JOIN EnterpriseAdmin_AMC.Procurement.PagosPlanificados PP
                ON SAACXP.NroUnico = PP.NroUnico
            WHERE SAACXP.TipoCxP = '10' 
               AND (SAACXP.NumeroD LIKE ?
               OR SACOMP.NumeroD LIKE ?
               OR SAPAGCXP.NumeroD LIKE ?
               OR SAPROV.Descrip LIKE ?)
               {date_filter}
            ORDER BY SAACXP.FechaE DESC
        """
        
        search_param = f"%{search}%"
        params = [search_param, search_param, search_param, search_param] + date_params
            
        cursor.execute(query, tuple(params))
        
        columns = [column[0] for column in cursor.description]
        results = []
        for row in cursor.fetchall():
            results.append(dict(zip(columns, row)))
            
        return {"data": results}
        
    except Exception as e:
        logging.error(f"Error fetching data: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

# --- REPORTS ENDPOINTS ---

@app.get("/api/reports/compras")
async def report_compras(desde: Optional[str] = Query(None), hasta: Optional[str] = Query(None)):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        date_filter = ""
        date_params = []
        if desde:
            date_filter += " AND CAST(SACOMP.FechaE AS DATE) >= ?"
            date_params.append(desde)
        if hasta:
            date_filter += " AND CAST(SACOMP.FechaE AS DATE) <= ?"
            date_params.append(hasta)
            
        if not desde and not hasta:
            date_filter = " AND SACOMP.FechaE >= DATE_SUB(CURDATE(), INTERVAL 1 YEAR)" if "MySQL" in database.DRIVER else " AND SACOMP.FechaE >= DATEADD(year, -1, GETDATE())"

        # Query for grouped stats in USD
        query = f"""
            SELECT
              SACOMP.Descrip AS Proveedor,
              SUM(SACOMP.MtoTotal / NULLIF(dt_emision.dolarbcv, 0)) AS TotalUSD,
              COUNT(SACOMP.NroUnico) AS CantidadFacturas
            FROM dbo.SACOMP WITH (NOLOCK)
            OUTER APPLY (
                SELECT TOP 1 dolarbcv 
                FROM EnterpriseAdmin_AMC.dbo.dolartoday WITH (NOLOCK)
                WHERE CAST(fecha AS DATE) <= CAST(SACOMP.FechaE AS DATE)
                ORDER BY fecha DESC
            ) dt_emision
            WHERE 1=1 {date_filter}
            GROUP BY SACOMP.Descrip
            ORDER BY TotalUSD DESC
        """
        cursor.execute(query, tuple(date_params))
        rows = cursor.fetchall()
        columns = [column[0] for column in cursor.description]
        data = [dict(zip(columns, row)) for row in rows]
        
        # Calculate Grand Total for percentages
        grand_total = sum(item['TotalUSD'] for item in data if item['TotalUSD'])
        for item in data:
            item['TotalUSD'] = float(item['TotalUSD']) if item['TotalUSD'] else 0
            item['Porcentaje'] = (item['TotalUSD'] / grand_total * 100) if grand_total > 0 else 0
            
        return {"data": data, "grand_total": float(grand_total) if grand_total else 0}
    except Exception as e:
        logging.error(f"Error in report_compras: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

# --- ENDPOINTS PROVEEDORES INDEXACION ---
@app.get("/api/procurement/providers")
async def get_provider_conditions():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Return all active providers from SAPROV and their conditions if any
        query = """
            SELECT p.CodProv, p.Descrip, p.activo, p.DiasCred AS SaprovDiasCred,
                   ISNULL(c.DiasNoIndexacion, 0) AS DiasNoIndexacion, 
                   ISNULL(c.BaseDiasCredito, 'EMISION') AS BaseDiasCredito, 
                   ISNULL(c.DiasVencimiento, p.DiasCred) AS DiasVencimiento,
                   ISNULL(c.ProntoPago1_Dias, 0) AS ProntoPago1_Dias, 
                   ISNULL(c.ProntoPago1_Pct, 0) AS ProntoPago1_Pct,
                   ISNULL(c.ProntoPago2_Dias, 0) AS ProntoPago2_Dias, 
                   ISNULL(c.ProntoPago2_Pct, 0) AS ProntoPago2_Pct,
                   COALESCE(c.Email, p.Email, '') AS Email
            FROM EnterpriseAdmin_AMC.dbo.SAPROV p WITH (NOLOCK)
            LEFT JOIN EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones c WITH (NOLOCK) ON p.CodProv = c.CodProv
            ORDER BY p.Descrip
        """
        cursor.execute(query)
        columns = [column[0] for column in cursor.description]
        results = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Convert Decimals to float for JSON serialization
        for r in results:
            r['ProntoPago1_Pct'] = float(r['ProntoPago1_Pct'])
            r['ProntoPago2_Pct'] = float(r['ProntoPago2_Pct'])
            
        return {"data": results}
    except Exception as e:
        logging.error(f"Error loading providers: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

@app.put("/api/procurement/providers/{cod_prov}")
async def update_provider_condition(cod_prov: str, payload: ProveedorCondicion):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Check if condition info exists
        check_query = "SELECT CodProv FROM EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones WHERE CodProv = ?"
        cursor.execute(check_query, (cod_prov,))
        if cursor.fetchone():
            update_query = """
                UPDATE EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones
                SET DiasNoIndexacion = ?, BaseDiasCredito = ?, DiasVencimiento = ?,
                    ProntoPago1_Dias = ?, ProntoPago1_Pct = ?,
                    ProntoPago2_Dias = ?, ProntoPago2_Pct = ?,
                    Email = ?
                WHERE CodProv = ?
            """
            cursor.execute(update_query, (
                payload.DiasNoIndexacion, payload.BaseDiasCredito, payload.DiasVencimiento,
                payload.ProntoPago1_Dias, payload.ProntoPago1_Pct,
                payload.ProntoPago2_Dias, payload.ProntoPago2_Pct,
                payload.Email,
                cod_prov
            ))
        else:
            insert_query = """
                INSERT INTO EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones 
                (CodProv, DiasNoIndexacion, BaseDiasCredito, DiasVencimiento, ProntoPago1_Dias, ProntoPago1_Pct, ProntoPago2_Dias, ProntoPago2_Pct, Email)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            cursor.execute(insert_query, (
                cod_prov, payload.DiasNoIndexacion, payload.BaseDiasCredito, payload.DiasVencimiento,
                payload.ProntoPago1_Dias, payload.ProntoPago1_Pct,
                payload.ProntoPago2_Dias, payload.ProntoPago2_Pct,
                payload.Email
            ))
            
        # Synchronize SAPROV native credit days
        update_saprov = "UPDATE EnterpriseAdmin_AMC.dbo.SAPROV SET DiasCred = ? WHERE CodProv = ?"
        cursor.execute(update_saprov, (payload.DiasVencimiento, cod_prov))
        
        conn.commit()
        return {"message": "Condiciones del proveedor actualizadas."}
    except Exception as e:
        logging.error(f"Error updating provider {cod_prov}: {e}", exc_info=True)
        if 'conn' in locals():
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

def enviar_correo_pago(destinatario: str, proveedor_nombre: str, nro_factura: str, pago_data, filepath: str):
    """Send payment notification via Gmail API.
    pago_data can be a single dict or a list of dicts (for multi-invoice payments).
    Supports multiple recipients separated by ';' in destinatario.
    """
    try:
        service = get_gmail_service()
        if not service:
            logging.warning("Gmail API no disponible. Ejecuta: .venv\\Scripts\\python.exe setup_gmail.py")
            return False
        
        emails = [e.strip() for e in destinatario.split(";") if e.strip()]
        if not emails:
            logging.warning(f"No valid email addresses found in: {destinatario}")
            return False
        
        # Normalize pago_data to list
        if isinstance(pago_data, dict):
            pagos_list = [pago_data]
        else:
            pagos_list = list(pago_data)
        
        is_multi = len(pagos_list) > 1
        
        remitente = os.getenv("SMTP_EMAIL", "")
        msg = MIMEMultipart()
        msg['From'] = remitente
        msg['To'] = ", ".join(emails)
        
        if is_multi:
            total_bs = sum(float(p.get('MontoBsAbonado', 0)) for p in pagos_list)
            msg['Subject'] = f"Soporte de Pago - {len(pagos_list)} Facturas - {proveedor_nombre}"
            facturas_str = ", ".join(p.get('NumeroD', '?') for p in pagos_list)
            cuerpo = f"""Estimados/as {proveedor_nombre},

Adjunto a este correo se encuentra el resumen de pagos correspondiente a las siguientes facturas: {facturas_str}.
El monto total pagado es de: Bs. {total_bs:,.2f}

Atentamente,
El equipo de Administracion."""
        else:
            monto_bs = pagos_list[0].get('MontoBsAbonado', 0)
            msg['Subject'] = f"Soporte de Pago y Resumen - Factura {nro_factura}"
            cuerpo = f"""Estimados/as {proveedor_nombre},

Adjunto a este correo se encuentra el soporte de pago correspondiente a la factura Nro: {nro_factura}.
El monto pagado total es de: Bs. {monto_bs:,.2f}

Atentamente,
El equipo de Administracion."""
        msg.attach(MIMEText(cuerpo, 'plain', 'utf-8'))
        
        # Generar Excel con formato profesional usando openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Resumen Pago'
        
        headers = ['Nro Factura', 'Referencia', 'Fecha de Pago', 'Monto Pagado (Bs)', 'Tasa de Cambio', 'Indexado', 'Monto Pagado (USD)']
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        data_font = Font(name='Calibri', size=11)
        money_fmt = '#,##0.00'
        rate_fmt = '#,##0.0000'
        
        for row_idx, p in enumerate(pagos_list, 2):
            ws.cell(row=row_idx, column=1, value=p.get('NumeroD', '')).font = data_font
            ws.cell(row=row_idx, column=2, value=p.get('Referencia', '')).font = data_font
            ws.cell(row=row_idx, column=3, value=p.get('FechaAbono', '')).font = data_font
            
            c_bs = ws.cell(row=row_idx, column=4, value=float(p.get('MontoBsAbonado', 0)))
            c_bs.font = data_font
            c_bs.number_format = money_fmt
            c_bs.alignment = Alignment(horizontal='right')
            
            c_tasa = ws.cell(row=row_idx, column=5, value=float(p.get('TasaCambioDiaAbono', 0)))
            c_tasa.font = data_font
            c_tasa.number_format = rate_fmt
            c_tasa.alignment = Alignment(horizontal='right')
            
            ws.cell(row=row_idx, column=6, value=p.get('AplicaIndexacion', 'No')).font = data_font
            
            c_usd = ws.cell(row=row_idx, column=7, value=float(p.get('MontoUsdAbonado', 0)))
            c_usd.font = data_font
            c_usd.number_format = money_fmt
            c_usd.alignment = Alignment(horizontal='right')
            
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).border = thin_border
        
        # Totals row
        if is_multi:
            tot_row = len(pagos_list) + 2
            total_font = Font(name='Calibri', bold=True, size=11)
            total_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
            
            ws.cell(row=tot_row, column=1, value='TOTALES').font = total_font
            ws.cell(row=tot_row, column=1).fill = total_fill
            for col in range(2, 4):
                ws.cell(row=tot_row, column=col).fill = total_fill
            
            c_total_bs = ws.cell(row=tot_row, column=4, value=sum(float(p.get('MontoBsAbonado', 0)) for p in pagos_list))
            c_total_bs.font = total_font
            c_total_bs.fill = total_fill
            c_total_bs.number_format = money_fmt
            c_total_bs.alignment = Alignment(horizontal='right')
            
            ws.cell(row=tot_row, column=5).fill = total_fill
            ws.cell(row=tot_row, column=6).fill = total_fill
            
            c_total_usd = ws.cell(row=tot_row, column=7, value=sum(float(p.get('MontoUsdAbonado', 0)) for p in pagos_list))
            c_total_usd.font = total_font
            c_total_usd.fill = total_fill
            c_total_usd.number_format = money_fmt
            c_total_usd.alignment = Alignment(horizontal='right')
            
            for col in range(1, 8):
                ws.cell(row=tot_row, column=col).border = thin_border
        
        # Auto-fit column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max(max_length + 3, 12)
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        part_excel = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part_excel.set_payload(excel_buffer.read())
        encoders.encode_base64(part_excel)
        fn_suffix = f"{len(pagos_list)}_Facturas" if is_multi else nro_factura
        part_excel.add_header("Content-Disposition", f"attachment; filename=Resumen_Pago_{fn_suffix}.xlsx")
        msg.attach(part_excel)
        
        if filepath and os.path.exists(filepath):
            with open(filepath, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
            encoders.encode_base64(part)
            ext = os.path.splitext(filepath)[1]
            part.add_header("Content-Disposition", f"attachment; filename=Comprobante_Pago_{fn_suffix}{ext}")
            msg.attach(part)
        
        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
        service.users().messages().send(userId='me', body={'raw': raw}).execute()
        logging.info(f"Correo enviado exitosamente a {', '.join(emails)} via Gmail API")
        return True
    except HttpError as e:
        logging.error(f"Gmail API HttpError al enviar correo a {destinatario}: {e}")
        return False
    except Exception as e:
        logging.error(f"Error al enviar correo a {destinatario}: {e}")
        return False

# Helper: wrap email send so a network error doesn't crash the payment transaction
def safe_send_email(destinatario: str, proveedor_nombre: str, nro_factura: str, pago_data, filepath: str) -> bool:
    try:
        return enviar_correo_pago(destinatario, proveedor_nombre, nro_factura, pago_data, filepath)
    except Exception as e:
        logging.warning(f"Email send failed (likely offline): {e}")
        return False

@app.post("/api/procurement/abonos-batch", response_model=None)
async def registrar_abonos_batch(
    pagos_json: str = Form(...),
    NotificarCorreo: bool = Form(False),
    archivo: UploadFile = File(None)
):
    """Register multiple payments in one transaction and send one consolidated email."""
    import json as _json
    try:
        pagos = _json.loads(pagos_json)
        if not isinstance(pagos, list) or len(pagos) == 0:
            raise HTTPException(status_code=400, detail="Se requiere al menos un pago.")
        
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        filepath = None
        if archivo and archivo.filename:
            ext = os.path.splitext(archivo.filename)[1]
            filename = f"{uuid.uuid4()}{ext}"
            filepath = f"static/uploads/{filename}"
            with open(filepath, "wb") as buffer:
                shutil.copyfileobj(archivo.file, buffer)
        
        for p in pagos:
            aplica_idx = 1 if p.get('AplicaIndexacion') in [True, 'true', 'True', 1] else 0
            cursor.execute("""
                INSERT INTO EnterpriseAdmin_AMC.dbo.CxP_Abonos 
                (NumeroD, CodProv, FechaAbono, MontoBsAbonado, TasaCambioDiaAbono, MontoUsdAbonado, AplicaIndexacion, Referencia, RutaComprobante, NotificarCorreo)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                p['NumeroD'], p['CodProv'], p['FechaAbono'],
                float(p.get('MontoBsAbonado', 0)), float(p.get('TasaCambioDiaAbono', 0)),
                float(p.get('MontoUsdAbonado', 0)), aplica_idx,
                p.get('Referencia', ''), filepath, 1 if NotificarCorreo else 0
            ))
        
        conn.commit()
        
        email_sent = False
        if NotificarCorreo:
            cod_prov = pagos[0]['CodProv']
            cursor.execute("""
                SELECT c.Email, p.Descrip 
                FROM EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones c 
                LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON c.CodProv = p.CodProv 
                WHERE c.CodProv = ?
            """, (cod_prov,))
            row = cursor.fetchone()
            if row and row.Email:
                pago_data_list = []
                for p in pagos:
                    pago_data_list.append({
                        "NumeroD": p['NumeroD'],
                        "CodProv": p['CodProv'],
                        "FechaAbono": p.get('FechaAbono', ''),
                        "MontoBsAbonado": float(p.get('MontoBsAbonado', 0)),
                        "MontoUsdAbonado": float(p.get('MontoUsdAbonado', 0)),
                        "TasaCambioDiaAbono": float(p.get('TasaCambioDiaAbono', 0)),
                        "AplicaIndexacion": "Sí" if p.get('AplicaIndexacion') in [True, 'true', 'True', 1] else "No",
                        "Referencia": p.get('Referencia', '')
                    })
                email_sent = safe_send_email(
                    row.Email, row.Descrip or "Proveedor",
                    f"{len(pagos)}_Facturas", pago_data_list, filepath
                )
        
        logging.info(f"Batch abonos: {len(pagos)} pagos registrados")
        return {"message": f"{len(pagos)} pagos registrados exitosamente.", "count": len(pagos), "email_sent": email_sent}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error batch abonos: {e}", exc_info=True)
        if 'conn' in locals():
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

@app.post("/api/procurement/abonos")
async def registrar_abono(
    NumeroD: str = Form(...),
    CodProv: str = Form(...),
    FechaAbono: str = Form(...),
    MontoBsAbonado: float = Form(...),
    TasaCambioDiaAbono: float = Form(...),
    MontoUsdAbonado: float = Form(...),
    AplicaIndexacion: str = Form(...),
    Referencia: str = Form(""),
    NotificarCorreo: bool = Form(False),
    force_send: bool = Form(False),
    archivo: UploadFile = File(None)
):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()

        aplica_idx = 1 if AplicaIndexacion.lower() == 'true' else 0
        notificar = 1 if NotificarCorreo else 0
        force = 1 if force_send else 0

        filepath = None
        if archivo and archivo.filename:
            ext = os.path.splitext(archivo.filename)[1]
            filename = f"{uuid.uuid4()}{ext}"
            filepath = f"static/uploads/{filename}"
            with open(filepath, "wb") as buffer:
                shutil.copyfileobj(archivo.file, buffer)

        insert_query = """
            INSERT INTO EnterpriseAdmin_AMC.dbo.CxP_Abonos 
            (NumeroD, CodProv, FechaAbono, MontoBsAbonado, TasaCambioDiaAbono, MontoUsdAbonado, AplicaIndexacion, Referencia, RutaComprobante, NotificarCorreo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        cursor.execute(insert_query, (
            NumeroD, CodProv, FechaAbono,
            MontoBsAbonado, TasaCambioDiaAbono,
            MontoUsdAbonado, aplica_idx, Referencia,
            filepath, notificar
        ))

        # Enviar correo si corresponde (notificar o force_send)
        if notificar == 1 or force == 1:
            cursor.execute("""
                SELECT c.Email, p.Descrip 
                FROM EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones c 
                LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON c.CodProv = p.CodProv 
                WHERE c.CodProv = ?
            """, (CodProv,))
            row = cursor.fetchone()
            if row and row.Email:
                pago_data = {
                    "NumeroD": NumeroD,
                    "CodProv": CodProv,
                    "FechaAbono": FechaAbono,
                    "MontoBsAbonado": MontoBsAbonado,
                    "MontoUsdAbonado": MontoUsdAbonado,
                    "TasaCambioDiaAbono": TasaCambioDiaAbono,
                    "AplicaIndexacion": "Sí" if AplicaIndexacion.lower() == 'true' else "No",
                    "Referencia": Referencia
                }
                email_sent = enviar_correo_pago(row.Email, row.Descrip or "Proveedor", NumeroD, pago_data, filepath)
                logging.info(f"Email sent flag: {email_sent}")
            else:
                logging.warning(f"No notification sent: missing email or provider logic cod={CodProv}")

        conn.commit()
        return {"message": "Abono registrado exitosamente.", "email_sent": (notificar == 1 or force == 1)}
    except Exception as e:
        logging.error(f"Error registering abono: {e}", exc_info=True)
        if 'conn' in locals():
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

# Helper: wrap email send so a network error doesn't crash the payment transaction
def safe_send_email(destinatario: str, proveedor_nombre: str, nro_factura: str, pago_data: dict, filepath: str) -> bool:
    try:
        return enviar_correo_pago(destinatario, proveedor_nombre, nro_factura, pago_data, filepath)
    except Exception as e:
        logging.warning(f"Email send failed (likely offline): {e}")
        return False

# Endpoint dedicated only to sending email (no payment insertion)
@app.post("/api/procurement/send-email", response_model=None)
async def send_email_only(
    NumeroD: str = Form(...),
    CodProv: str = Form(...),
    archivo: UploadFile = File(None)
):
    """Send a payment notification email for an existing invoice WITHOUT inserting a new payment record."""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT c.Email, p.Descrip 
            FROM EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones c 
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON c.CodProv = p.CodProv 
            WHERE c.CodProv = ?
        """, (CodProv,))
        row = cursor.fetchone()

        if not row or not row.Email:
            return {"email_sent": False, "message": "Proveedor sin email configurado."}

        # Save attachment if provided
        filepath = None
        if archivo and archivo.filename:
            ext = os.path.splitext(archivo.filename)[1]
            filename = f"{uuid.uuid4()}{ext}"
            filepath = f"static/uploads/{filename}"
            with open(filepath, "wb") as buffer:
                shutil.copyfileobj(archivo.file, buffer)

        # Fetch last abono for context
        cursor.execute("""
            SELECT TOP 1 FechaAbono, MontoBsAbonado, MontoUsdAbonado, TasaCambioDiaAbono, AplicaIndexacion, Referencia, RutaComprobante
            FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos
            WHERE NumeroD = ? AND CodProv = ?
            ORDER BY FechaAbono DESC
        """, (NumeroD, CodProv))
        abono = cursor.fetchone()

        if abono:
            pago_data = {
                "NumeroD": NumeroD,
                "CodProv": CodProv,
                "FechaAbono": str(abono.FechaAbono).split(" ")[0] if abono.FechaAbono else "-", # type: ignore
                "MontoBsAbonado": abono.MontoBsAbonado or 0,
                "MontoUsdAbonado": abono.MontoUsdAbonado or 0,
                "TasaCambioDiaAbono": abono.TasaCambioDiaAbono or 0,
                "AplicaIndexacion": "Sí" if abono.AplicaIndexacion else "No",
                "Referencia": abono.Referencia or "Re-envio de soporte"
            }
            # Use the original comprobante if no new file was uploaded
            attach = filepath or (str(abono.RutaComprobante) if abono.RutaComprobante else None)
        else:
            pago_data = {
                "NumeroD": NumeroD, "CodProv": CodProv,
                "FechaAbono": "-", "MontoBsAbonado": 0,
                "MontoUsdAbonado": 0, "TasaCambioDiaAbono": 0,
                "AplicaIndexacion": "No", "Referencia": "Re-envio de soporte"
            }
            attach = filepath

        sent = safe_send_email(row.Email, row.Descrip or "Proveedor", NumeroD, pago_data, attach) # type: ignore
        return {"email_sent": sent, "message": "Correo enviado." if sent else "No se pudo enviar el correo."}
    except Exception as e:
        logging.error(f"Error send-email {NumeroD}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

# PATCH endpoint to edit invoice fields - syncs all Saint tables (SAACXP + SACOMP + SAPROV)
@app.patch("/api/cuentas-por-pagar/{numeroD}", response_model=None)
async def editar_factura(numeroD: str, cod_prov: str = Query(default=''), payload: dict = Body(...)):
    """
    Update invoice fields and keep Saint tables fully aligned.
    Fields: FechaE, FechaI, FechaV, SaldoAct (maps to SAACXP.Saldo + SAPROV.Saldo delta), MontoFacturaBS, MontoFacturaUSD, Notas10.
    cod_prov query param is required to correctly identify the SACOMP row.
    """
    print(f"\n[PATCH-DEV] numeroD={numeroD} cod_prov={cod_prov} payload={payload}")
    # Write to a persistent debug file with absolute path
    log_file = r"C:\source\cuentasporpagarDev\patch_debug.log"
    try:
        with open(log_file, "a", encoding="utf-8") as f:
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            f.write(f"[{timestamp}] DEV numeroD={numeroD} cod_prov={cod_prov} payload={payload}\n")
    except Exception as e:
        print(f"Error writing to log: {e}")

    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()

        # Update cod_prov from payload if provided
        if "CodProv" in payload:
            cod_prov = payload["CodProv"]

        # ── 1. SAACXP.Saldo + SAPROV.Saldo (sync using delta) ────────────────
        if "SaldoAct" in payload and payload["SaldoAct"] is not None:
            nuevo_saldo = float(payload["SaldoAct"])
            
            # Fetch previous Saldo to calculate delta
            cursor.execute(
                "SELECT Saldo, CodProv FROM EnterpriseAdmin_AMC.dbo.SAACXP WHERE NumeroD = ? AND TipoCxP = '10'", 
                (numeroD,)
            )
            row = cursor.fetchone()
            
            if row is not None:
                viejo_saldo = float(row.Saldo) if row.Saldo is not None else 0.0
                delta = nuevo_saldo - viejo_saldo
                # Only use cod_prov from DB if we don't have it from payload
                if not cod_prov:
                    cod_prov = row.CodProv
                
                # Update SAACXP.Saldo
                cursor.execute(
                    """UPDATE EnterpriseAdmin_AMC.dbo.SAACXP
                       SET Saldo = ?
                       WHERE NumeroD = ? AND TipoCxP = '10'""",
                    (nuevo_saldo, numeroD)
                )
                
                # Update SAPROV.Saldo by applying the delta
                prov_for_delta = cod_prov or row.CodProv
                if delta != 0 and prov_for_delta:
                    cursor.execute(
                        """UPDATE EnterpriseAdmin_AMC.dbo.SAPROV
                           SET Saldo = Saldo + ?
                           WHERE CodProv = ?""",
                        (delta, prov_for_delta)
                    )

        # ── 2. SAACXP: Monto si cambia MontoFacturaBS ─────────────────────────
        if "MontoFacturaBS" in payload and payload["MontoFacturaBS"] is not None:
            cursor.execute(
                """UPDATE EnterpriseAdmin_AMC.dbo.SAACXP
                   SET Monto = ?
                   WHERE NumeroD = ? AND TipoCxP = '10'""",
                (float(payload["MontoFacturaBS"]), numeroD)
            )

        # ── 3. SACOMP: fechas + montos ────────────────────────────────────────
        comp_fields = []
        comp_params = []
        if "FechaE" in payload and payload["FechaE"]:
            comp_fields.append("FechaE = ?")
            comp_params.append(payload["FechaE"])
            cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SAACXP SET FechaE = ? WHERE NumeroD = ? AND TipoCxP = '10'", (payload["FechaE"], numeroD))
        if "FechaI" in payload and payload["FechaI"]:
            comp_fields.append("FechaI = ?")
            comp_params.append(payload["FechaI"])
            cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SAACXP SET FechaI = ? WHERE NumeroD = ? AND TipoCxP = '10'", (payload["FechaI"], numeroD))
        if "FechaV" in payload and payload["FechaV"]:
            comp_fields.append("FechaV = ?")
            comp_params.append(payload["FechaV"])
            cursor.execute("UPDATE EnterpriseAdmin_AMC.dbo.SAACXP SET FechaV = ? WHERE NumeroD = ? AND TipoCxP = '10'", (payload["FechaV"], numeroD))
            
        if "MontoFacturaBS" in payload and payload["MontoFacturaBS"] is not None:
            monto_bs = float(payload["MontoFacturaBS"])
            comp_fields.append("Credito = ?")
            comp_params.append(monto_bs)
            comp_fields.append("MtoTotal = ?")     # keep MtoTotal aligned
            comp_params.append(monto_bs)

        if "Notas10" in payload:
            notas10_val = str(payload["Notas10"]).strip() if payload["Notas10"] is not None else ""
            if notas10_val == "1":
                comp_fields.append("Notas10 = ?")
                comp_params.append("1")
            elif notas10_val == "0" or notas10_val == "":
                # '0' or 'Sin cambio' means clear the field
                comp_fields.append("Notas10 = NULL")
        
        if comp_fields:
            set_clause = ", ".join(comp_fields)
            if cod_prov:
                comp_params.append(cod_prov)
                comp_params.append(numeroD)
                cursor.execute(
                    f"UPDATE EnterpriseAdmin_AMC.dbo.SACOMP SET {set_clause} WHERE CodProv = ? AND NumeroD = ?",
                    tuple(comp_params)
                )
            else:
                comp_params.append(numeroD) # type: ignore
                cursor.execute(
                    f"UPDATE EnterpriseAdmin_AMC.dbo.SACOMP SET {set_clause} WHERE NumeroD = ?",
                    tuple(comp_params)
                )

            print(f"[PATCH-DEV SACOMP] Affected Rows: {cursor.rowcount}")
            with open(log_file, "a", encoding="utf-8") as f:
                f.write(f"[PATCH-DEV SACOMP] Affected Rows: {cursor.rowcount}\n")

        conn.commit()
        return {"message": f"Factura {numeroD} actualizada y tablas Saint sincronizadas correctamente."}
    except Exception as e:
        logging.error(f"Error updating factura {numeroD}: {e}", exc_info=True)
        if 'conn' in locals():
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/procurement/cxp-status")
async def get_cxp_status(cod_prov: str = Query(...), numero_d: str = Query(...)):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # We need to fetch:
        # 1. Invoice details from SACOMP/SAACXP (Monto, Saldo, Fechas, Tasa original)
        # 2. Provider conditions from ProveedorCondiciones
        # 3. Sum of all Abonos from CxP_Abonos
        
        query = """
            SELECT 
                cxp.NumeroD, cxp.CodProv, cxp.Monto, cxp.Saldo, 
                cxp.FechaE, cxp.FechaV AS FechaVSaint,
                comp.FechaI, comp.Notas10,
                ISNULL(cond.DiasNoIndexacion, 0) AS DiasNoIndexacion,
                ISNULL(cond.BaseDiasCredito, 'EMISION') AS BaseDiasCredito,
                ISNULL(cond.DiasVencimiento, prov.diascred) AS DiasVencimiento,
                ISNULL(cond.ProntoPago1_Dias, 0) AS ProntoPago1_Dias,
                ISNULL(cond.ProntoPago1_Pct, 0) AS ProntoPago1_Pct,
                ISNULL(cond.ProntoPago2_Dias, 0) AS ProntoPago2_Dias,
                ISNULL(cond.ProntoPago2_Pct, 0) AS ProntoPago2_Pct,
                prov.Descrip AS ProveedorNombre,
                prov.NumeroUP, prov.FechaUP, prov.MontoUP,
                dt_emision.dolarbcv AS TasaEmision,
                dt_actual.dolarbcv AS TasaActual,
                ISNULL(abonos.TotalUsdAbonado, 0) AS TotalUsdAbonado,
                ISNULL(abonos.TotalBsAbonado, 0) AS TotalBsAbonado
            FROM dbo.SAACXP cxp
            LEFT JOIN dbo.SACOMP comp ON cxp.CodProv = comp.CodProv AND cxp.NumeroD = comp.NumeroD
            LEFT JOIN dbo.SAPROV prov ON cxp.CodProv = prov.CodProv
            LEFT JOIN EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones cond ON cxp.CodProv = cond.CodProv
            OUTER APPLY (
                SELECT SUM(MontoUsdAbonado) as TotalUsdAbonado, SUM(MontoBsAbonado) as TotalBsAbonado
                FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos
                WHERE CodProv = cxp.CodProv AND NumeroD = cxp.NumeroD
            ) abonos
            OUTER APPLY (
                SELECT TOP 1 dolarbcv 
                FROM EnterpriseAdmin_AMC.dbo.dolartoday 
                WHERE CAST(fecha AS DATE) <= CAST(cxp.FechaE AS DATE)
                ORDER BY fecha DESC
            ) dt_emision
            OUTER APPLY (
                SELECT TOP 1 dolarbcv 
                FROM EnterpriseAdmin_AMC.dbo.dolartoday 
                WHERE dolarbcv IS NOT NULL
                ORDER BY id DESC
            ) dt_actual
            WHERE cxp.CodProv = ? AND cxp.NumeroD = ? AND cxp.TipoCxP = '10'
        """
        cursor.execute(query, (cod_prov, numero_d))
        row = cursor.fetchone()
        
        if not row:
            raise HTTPException(status_code=404, detail="Factura no encontrada")
            
        columns = [column[0] for column in cursor.description]
        data = dict(zip(columns, row))
        
        # Calculate dynamic dates
        from datetime import datetime, timedelta
        
        # Base date for calculations
        base_date_str = data['FechaE'] if data['BaseDiasCredito'] == 'EMISION' else (data['FechaI'] or data['FechaE'])
        # base_date_str comes as string or datetime from DB (pyodbc parses to datetime)
        if isinstance(base_date_str, str):
            base_date = datetime.strptime(base_date_str[:10], '%Y-%m-%d')
        else:
            base_date = base_date_str
            
        if data.get('Notas10') == '1':
            fecha_ni = base_date
        else:
            fecha_ni = base_date + timedelta(days=int(data['DiasNoIndexacion']))
            
        fecha_v = base_date + timedelta(days=int(data['DiasVencimiento']))
        fecha_pp1 = base_date + timedelta(days=int(data['ProntoPago1_Dias']))
        fecha_pp2 = base_date + timedelta(days=int(data['ProntoPago2_Dias']))
        
        data['FechaNI_Calculada'] = fecha_ni.strftime('%Y-%m-%d')
        data['FechaV_Calculada'] = fecha_v.strftime('%Y-%m-%d')
        data['FechaPP1'] = fecha_pp1.strftime('%Y-%m-%d')
        data['FechaPP2'] = fecha_pp2.strftime('%Y-%m-%d')
        
        # Convert numeric types to float for JSON
        for k, v in data.items():
            if hasattr(v, 'quantize') or isinstance(v, float): # Decimal or float
                data[k] = float(v) if v is not None else 0.0
            elif isinstance(v, datetime): # type: ignore
                data[k] = v.strftime('%Y-%m-%d')
                
        # Compute specific fields
        monto_original_usd = data['Monto'] / data['TasaEmision'] if data['TasaEmision'] else 0
        saldo_usd = monto_original_usd - data['TotalUsdAbonado']
        data['MontoOriginalUSD'] = round(monto_original_usd, 2)
        data['SaldoRestanteUSD'] = round(saldo_usd, 2)
        # Note: DiferenciaBs is typically calculated on the frontend before closing an invoice 
        # based on exactly what was paid vs what the invoice was worth originally,
        # but the backend provides all necessary totals.

        # Fetch Abonos History
        history_query = """
            SELECT FechaAbono, MontoBsAbonado, TasaCambioDiaAbono, MontoUsdAbonado, AplicaIndexacion, Referencia
            FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos
            WHERE CodProv = ? AND NumeroD = ?
            ORDER BY FechaAbono ASC
        """
        cursor.execute(history_query, (cod_prov, numero_d))
        history_cols = [column[0] for column in cursor.description]
        history_data = [dict(zip(history_cols, row)) for row in cursor.fetchall()]
        
        for record in history_data:
            for k, v in record.items():
                if hasattr(v, 'quantize') or isinstance(v, float):
                    record[k] = float(v)
                elif isinstance(v, datetime): # type: ignore
                    record[k] = v.strftime('%Y-%m-%d')
                    
        data['HistorialAbonos'] = history_data

        return {"data": data}
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error fetching cxp status: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/procurement/debit-notes")
async def get_debit_notes():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # We need invoices where Total amount paid (Bs) > original invoice amount (Bs)
        # We join Cxp_Abonos and SAACXP.
        query = """
            SELECT 
                cxp.CodProv, 
                prov.Descrip AS ProveedorNombre,
                cxp.NumeroD, 
                cxp.FechaE AS FechaEmision,
                cxp.Monto AS MontoOriginalBs,
                ISNULL(abonos.TotalBsAbonado, 0) AS TotalBsAbonado,
                ISNULL(abonos.TotalBsAbonado, 0) - cxp.Monto AS MontoNotaDebitoBs,
                ISNULL(dnt.Estatus, 'PENDIENTE') AS Estatus,
                dnt.NotaDebitoID,
                dnt.MontoRetencionBs AS DB_MontoRetencionBs,
                COALESCE(cond.Email, prov.Email, '') AS Email,
                ISNULL(prov.PorctRet, 0) AS PorctRet
            FROM dbo.SAACXP cxp
            INNER JOIN dbo.SAPROV prov ON cxp.CodProv = prov.CodProv
            LEFT JOIN EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones cond ON cxp.CodProv = cond.CodProv
            LEFT JOIN EnterpriseAdmin_AMC.Procurement.DebitNotesTracking dnt ON cxp.CodProv = dnt.CodProv AND cxp.NumeroD = dnt.NumeroD
            CROSS APPLY (
                SELECT SUM(a.MontoBsAbonado) as TotalBsAbonado
                FROM EnterpriseAdmin_AMC.dbo.CxP_Abonos a
                WHERE a.CodProv = cxp.CodProv AND a.NumeroD = cxp.NumeroD
            ) abonos
            WHERE cxp.TipoCxP = '10' 
              AND ISNULL(abonos.TotalBsAbonado, 0) > cxp.Monto + 0.1
            ORDER BY prov.Descrip ASC, cxp.FechaE DESC
        """
        cursor.execute(query)
        columns = [column[0] for column in cursor.description]
        results = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Convert Decimals
        for r in results:
            r['MontoOriginalBs'] = float(r['MontoOriginalBs'] or 0)
            r['TotalBsAbonado'] = float(r['TotalBsAbonado'] or 0)
            r['MontoNotaDebitoBs'] = float(r['MontoNotaDebitoBs'] or 0)
            r['PorctRet'] = float(r['PorctRet'] or 0)
            
            if r['Estatus'] == 'EMITIDA' and r['DB_MontoRetencionBs'] is not None:
                r['MontoRetencionBs'] = float(r['DB_MontoRetencionBs'])
            else:
                # Estimar Retención (usando 16% base IVA)
                r['MontoRetencionBs'] = float(r['MontoNotaDebitoBs'] * 0.16 * (r['PorctRet'] / 100.0))
            
            r.pop('DB_MontoRetencionBs', None)
            
        return {"data": results}
    except Exception as e:
        logging.error(f"Error fetching debit notes: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

@app.post("/api/procurement/debit-notes/send-request")
async def send_debit_note_requests(payload: DebitNoteActionRequest):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # In a real-world scenario, here we would:
        # 1. Group payload.Invoices by CodProv.
        # 2. Iterate each provider, fetch their Email.
        # 3. Generate an Excel attachment /api/export/debit-notes?cod_prov=...
        # 4. Use python standard smtplib using os.getenv("SMTP_USERNAME") and os.getenv("SMTP_PASSWORD")
        # 5. Send email. 
        # For now, we will mark them as SOLICITUD_ENVIADA to complete the process tracking for the user.
        
        for inv in payload.Invoices:
            # Check if exists in tracking table
            cursor.execute("SELECT CodProv FROM EnterpriseAdmin_AMC.Procurement.DebitNotesTracking WHERE CodProv=? AND NumeroD=?", (inv.CodProv, inv.NumeroD))
            if cursor.fetchone():
                cursor.execute("UPDATE EnterpriseAdmin_AMC.Procurement.DebitNotesTracking SET Estatus='SOLICITUD_ENVIADA', FechaSolicitud=GETDATE() WHERE CodProv=? AND NumeroD=?", (inv.CodProv, inv.NumeroD))
            else:
                cursor.execute("INSERT INTO EnterpriseAdmin_AMC.Procurement.DebitNotesTracking (CodProv, NumeroD, Estatus, FechaSolicitud) VALUES (?, ?, 'SOLICITUD_ENVIADA', GETDATE())", (inv.CodProv, inv.NumeroD))
        
        conn.commit()
        return {"message": f"Se han marcado {len(payload.Invoices)} facturas como Solicitud Enviada."}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        logging.error(f"Error send_debit_note_requests: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@app.post("/api/procurement/debit-notes/register")
async def register_debit_note(payload: DebitNoteRegisterRequest):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        for inv in payload.Invoices:
            cursor.execute("SELECT CodProv FROM EnterpriseAdmin_AMC.Procurement.DebitNotesTracking WHERE CodProv=? AND NumeroD=?", (inv.CodProv, inv.NumeroD))
            if cursor.fetchone():
                cursor.execute("UPDATE EnterpriseAdmin_AMC.Procurement.DebitNotesTracking SET Estatus='EMITIDA', NotaDebitoID=?, MontoRetencionBs=?, FechaEmision=GETDATE() WHERE CodProv=? AND NumeroD=?", (payload.NotaDebitoID, inv.MontoRetencionBs, inv.CodProv, inv.NumeroD))
            else:
                cursor.execute("INSERT INTO EnterpriseAdmin_AMC.Procurement.DebitNotesTracking (CodProv, NumeroD, Estatus, NotaDebitoID, MontoRetencionBs, FechaEmision) VALUES (?, ?, 'EMITIDA', ?, ?, GETDATE())", (inv.CodProv, inv.NumeroD, payload.NotaDebitoID, inv.MontoRetencionBs))
        
        conn.commit()
        return {"message": "Notas de débito registradas correctamente."}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        logging.error(f"Error register_debit_note: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@app.get("/api/exchange-rate")
async def get_exchange_rate(fecha: str = Query(...)):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        query = """
            SELECT TOP 1 dolarbcv 
            FROM EnterpriseAdmin_AMC.dbo.dolartoday 
            WHERE CAST(fecha AS DATE) <= CAST(? AS DATE)
            ORDER BY fecha DESC
        """
        cursor.execute(query, (fecha,))
        row = cursor.fetchone()
        if not row:
            return {"rate": None}
        return {"rate": float(row[0])}
    except Exception as e:
        logging.error(f"Error fetching exchange rate: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/reports/aging")
async def report_aging():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        query = """
            SELECT
              SAPROV.CodProv,
              SAPROV.Descrip AS Proveedor,
              SUM(CASE WHEN DATEDIFF(day, SAACXP.FechaV, GETDATE()) <= 0 THEN SAACXP.Saldo / NULLIF(dt_emision.dolarbcv, 0) ELSE 0 END) AS PorVencer,
              SUM(CASE WHEN DATEDIFF(day, SAACXP.FechaV, GETDATE()) BETWEEN 1 AND 30 THEN SAACXP.Saldo / NULLIF(dt_emision.dolarbcv, 0) ELSE 0 END) AS Dias_1_30,
              SUM(CASE WHEN DATEDIFF(day, SAACXP.FechaV, GETDATE()) BETWEEN 31 AND 60 THEN SAACXP.Saldo / NULLIF(dt_emision.dolarbcv, 0) ELSE 0 END) AS Dias_31_60,
              SUM(CASE WHEN DATEDIFF(day, SAACXP.FechaV, GETDATE()) BETWEEN 61 AND 90 THEN SAACXP.Saldo / NULLIF(dt_emision.dolarbcv, 0) ELSE 0 END) AS Dias_61_90,
              SUM(CASE WHEN DATEDIFF(day, SAACXP.FechaV, GETDATE()) > 90 THEN SAACXP.Saldo / NULLIF(dt_emision.dolarbcv, 0) ELSE 0 END) AS Mas_90,
              SUM(SAACXP.Saldo / NULLIF(dt_emision.dolarbcv, 0)) AS Total
            FROM dbo.SAACXP WITH (NOLOCK)
            LEFT JOIN dbo.SAPROV WITH (NOLOCK) ON SAACXP.CodProv = SAPROV.CodProv
            OUTER APPLY (
                SELECT TOP 1 dolarbcv 
                FROM EnterpriseAdmin_AMC.dbo.dolartoday WITH (NOLOCK)
                WHERE CAST(fecha AS DATE) <= CAST(SAACXP.FechaE AS DATE)
                ORDER BY fecha DESC
            ) dt_emision
            WHERE SAACXP.Saldo > 0 AND SAACXP.TipoCxP = '10'
            GROUP BY SAPROV.CodProv, SAPROV.Descrip
            ORDER BY Total DESC
        """
        cursor.execute(query)
        columns = [column[0] for column in cursor.description]
        data = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Ensure floating point for JSON serialization
        for item in data:
            for k in ['PorVencer', 'Dias_1_30', 'Dias_31_60', 'Dias_61_90', 'Mas_90', 'Total']:
                item[k] = float(item[k]) if item[k] else 0
                
        return {"data": data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@app.get("/api/reports/cashflow")
async def report_cashflow(desde: str = None, hasta: str = None):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        date_filter_pagos = ""
        date_filter_gastos = ""
        params_pagos = []
        params_gastos = []

        if desde:
            date_filter_pagos += "CAST(COALESCE(PP.FechaPlanificada, SAACXP.FechaV) AS DATE) >= ?"
            date_filter_gastos += "CAST(fecha_proyectada AS DATE) >= ?"
            params_pagos.append(desde)
            params_gastos.append(desde)
            
        if hasta:
            if desde:
                date_filter_pagos += " AND "
                date_filter_gastos += " AND "
            date_filter_pagos += "CAST(COALESCE(PP.FechaPlanificada, SAACXP.FechaV) AS DATE) <= ?"
            date_filter_gastos += "CAST(fecha_proyectada AS DATE) <= ?"
            params_pagos.append(hasta)
            params_gastos.append(hasta)
            
        if not date_filter_pagos:
            date_filter_pagos = "1=1"
        if not date_filter_gastos:
            date_filter_gastos = "1=1"

        params = params_pagos + params_gastos + params_gastos

        query = f"""
            WITH Facturas AS (
                SELECT 
                   CAST(COALESCE(PP.FechaPlanificada, SAACXP.FechaV) AS DATE) AS Fecha,
                   SUM(SAACXP.Saldo) AS SaldoProyectado,
                   SUM(SAACXP.Saldo / NULLIF(dt_emision.dolarbcv, 0)) AS SaldoProyectadoUSD
                FROM dbo.SAACXP WITH (NOLOCK)
                LEFT JOIN EnterpriseAdmin_AMC.Procurement.PagosPlanificados PP ON SAACXP.NroUnico = PP.NroUnico
                OUTER APPLY (
                    SELECT TOP 1 dolarbcv 
                    FROM EnterpriseAdmin_AMC.dbo.dolartoday 
                    WHERE CAST(fecha AS DATE) <= CAST(SAACXP.FechaE AS DATE)
                    ORDER BY fecha DESC
                ) dt_emision
                WHERE SAACXP.Saldo > 0 AND SAACXP.TipoCxP = '10' AND {date_filter_pagos}
                GROUP BY CAST(COALESCE(PP.FechaPlanificada, SAACXP.FechaV) AS DATE)
            ),
            GastosFijos AS (
                SELECT 
                    CAST(fecha_proyectada AS DATE) AS Fecha,
                    SUM(monto_usd) AS SalidasFijosUSD
                FROM EnterpriseAdmin_AMC.Procurement.GastosProgramados WITH (NOLOCK)
                WHERE tipo = 'Farmacia' AND estado = 'Pendiente' AND {date_filter_gastos}
                GROUP BY CAST(fecha_proyectada AS DATE)
            ),
            GastosPersonales AS (
                SELECT 
                    CAST(fecha_proyectada AS DATE) AS Fecha,
                    SUM(monto_usd) AS SalidasPersonalesUSD
                FROM EnterpriseAdmin_AMC.Procurement.GastosProgramados WITH (NOLOCK)
                WHERE tipo = 'Personal' AND estado = 'Pendiente' AND {date_filter_gastos}
                GROUP BY CAST(fecha_proyectada AS DATE)
            ),
            Fechas AS (
                SELECT Fecha FROM Facturas
                UNION
                SELECT Fecha FROM GastosFijos
                UNION
                SELECT Fecha FROM GastosPersonales
            )
            SELECT 
                FORMAT(F.Fecha, 'yyyy-MM-dd') AS Periodo,
                CAST(ISNULL(ROUND(FCT.SaldoProyectado, 2), 0) AS FLOAT) AS SaldoProyectado,
                CAST(ISNULL(FCT.SaldoProyectadoUSD, 0) AS FLOAT) AS FacturasUSD,
                CAST(ISNULL(GF.SalidasFijosUSD, 0) AS FLOAT) AS GastosFijosUSD,
                CAST(ISNULL(GP.SalidasPersonalesUSD, 0) AS FLOAT) AS GastosPersonalesUSD,
                CAST(ISNULL(FCT.SaldoProyectadoUSD, 0) + ISNULL(GF.SalidasFijosUSD, 0) + ISNULL(GP.SalidasPersonalesUSD, 0) AS FLOAT) AS SaldoProyectadoUSD
            FROM Fechas F
            LEFT JOIN Facturas FCT ON F.Fecha = FCT.Fecha
            LEFT JOIN GastosFijos GF ON F.Fecha = GF.Fecha
            LEFT JOIN GastosPersonales GP ON F.Fecha = GP.Fecha
            ORDER BY F.Fecha
        """
        cursor.execute(query, params)
        columns = [column[0] for column in cursor.description]
        return {"data": [dict(zip(columns, row)) for row in cursor.fetchall()]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()


@app.get("/api/exchange-rate")
async def get_exchange_rate():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT TOP 1 dolarbcv FROM EnterpriseAdmin_AMC.dolartoday ORDER BY Fecha DESC")
        row = cursor.fetchone()
        rate = row[0] if row else 0
        return {"rate": rate}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@app.get("/api/reports/dpo")
async def report_dpo():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        query = """
            SELECT 
                FORMAT(SAACXP.FechaE, 'yyyy-MM') AS Periodo,
                AVG(DATEDIFF(day, SAACXP.FechaE, SAPAGCXP.FechaE)) AS PromedioDiasPago,
                COUNT(SAACXP.NroUnico) AS FacturasPagadas
            FROM dbo.SAACXP
            INNER JOIN dbo.SAPAGCXP ON SAACXP.NroUnico = SAPAGCXP.NroUnico
            WHERE SAACXP.Saldo <= 0 AND SAACXP.TipoCxP = '10' AND SAPAGCXP.FechaE >= DATEADD(year, -1, GETDATE())
            GROUP BY FORMAT(SAACXP.FechaE, 'yyyy-MM')
            ORDER BY Periodo
        """
        cursor.execute(query)
        columns = [column[0] for column in cursor.description]
        return {"data": [dict(zip(columns, row)) for row in cursor.fetchall()]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

# --- FORECAST & EVENTS ENDPOINTS ---

@app.get("/api/reports/forecast-sales")
async def report_forecast_sales(desde: str = None, hasta: str = None):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        where_ventas = ""
        where_reales = ""
        params_ventas = []
        params_reales = []
        
        if desde:
            where_ventas += " AND CAST(fecha_proyeccion AS DATE) >= ?"
            where_reales += " AND CAST(fecha AS DATE) >= ?"
            params_ventas.append(desde)
            params_reales.append(desde)
        else:
            where_ventas += " AND CAST(fecha_proyeccion AS DATE) >= CAST(GETDATE() AS DATE)"
            where_reales += " AND CAST(fecha AS DATE) >= DATEADD(day, -30, CAST(GETDATE() AS DATE))"
            
        if hasta:
            where_ventas += " AND CAST(fecha_proyeccion AS DATE) <= ?"
            where_reales += " AND CAST(fecha AS DATE) <= ?"
            params_ventas.append(hasta)
            params_reales.append(hasta)
            
        where_reales += " AND CAST(fecha AS DATE) < CAST(GETDATE() AS DATE)"
        where_ventas += " AND CAST(fecha_proyeccion AS DATE) >= CAST(GETDATE() AS DATE)"

        query = f"""
            SELECT 
                FORMAT(CAST(fecha_proyeccion AS DATE), 'yyyy-MM-dd') AS Periodo,
                CAST(monto_proyectado_ves AS FLOAT) AS VentasProyectadas,
                CAST(monto_proyectado_usd AS FLOAT) AS VentasProyectadasUSD
            FROM EnterpriseAdmin_AMC.Procurement.sales_forecast WITH (NOLOCK)
            WHERE 1=1 {where_ventas}
            UNION ALL
            SELECT 
                FORMAT(CAST(fecha AS DATE), 'yyyy-MM-dd') AS Periodo,
                CAST(SUM(MtoVentas) AS FLOAT) AS VentasProyectadas,
                CAST(SUM(CAST(MtoDolar AS FLOAT)) AS FLOAT) AS VentasProyectadasUSD
            FROM EnterpriseAdmin_AMC.dbo.CUSTOM_SAEVTA WITH (NOLOCK)
            WHERE 1=1 {where_reales}
            GROUP BY CAST(fecha AS DATE), FORMAT(CAST(fecha AS DATE), 'yyyy-MM-dd')
            ORDER BY Periodo
        """
        params = params_ventas + params_reales
        
        cursor.execute(query, params)
        columns = [column[0] for column in cursor.description]
        return {"data": [dict(zip(columns, row)) for row in cursor.fetchall()]}
    except Exception as e:
        logging.error(f"Error in forecast-sales: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@app.get("/api/reports/forecast-consolidated")
async def report_forecast_consolidated(
    desde: str = None, 
    hasta: str = None,
    fecha_arranque: str = None,
    caja_usd: float = 0.0,
    caja_bs: float = 0.0,
    delay_days: int = 1
):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # If no start date provided, use today
        if not fecha_arranque:
            from datetime import date
            fecha_arranque = date.today().isoformat()
            
        p_pagos = [fecha_arranque]
        p_ventas = [fecha_arranque]
        p_ventas_real = [fecha_arranque]
        p_gastos = [fecha_arranque]

        # Everything is calculated ALWAYS from fecha_arranque (Day Zero).
        # We fetch all history from Day Zero to compute the running total properly.
        # The frontend will just "hide" rows using JS if `desde` > `fecha_arranque`, 
        # but the backend must compute everything from Day Zero.
        date_filter_pagos = "CAST(COALESCE(PP.FechaPlanificada, SAACXP.FechaV) AS DATE) >= ?"
        date_filter_ventas = "CAST(DATEADD(day, ?, fecha_proyeccion) AS DATE) > ?"
        date_filter_ventas_real = "CAST(DATEADD(day, ?, fecha) AS DATE) > ?"
        date_filter_gastos = "CAST(fecha_proyectada AS DATE) >= ?"
        
        # Insert delay_days before fecha_arranque for Ventas DateAdd
        p_ventas.insert(0, delay_days)
        p_ventas_real.insert(0, delay_days)
        
        if hasta:
            date_filter_pagos += " AND CAST(COALESCE(PP.FechaPlanificada, SAACXP.FechaV) AS DATE) <= ?"
            date_filter_ventas += " AND CAST(DATEADD(day, ?, fecha_proyeccion) AS DATE) <= ?"
            date_filter_ventas_real += " AND CAST(DATEADD(day, ?, fecha) AS DATE) <= ?"
            date_filter_gastos += " AND CAST(fecha_proyectada AS DATE) <= ?"
            p_pagos.append(hasta)
            p_ventas.extend([delay_days, hasta])
            p_ventas_real.extend([delay_days, hasta])
            p_gastos.append(hasta)
            
        # We only want REAL sales for past dates
        date_filter_ventas_real += " AND CAST(fecha AS DATE) < CAST(GETDATE() AS DATE)"
        # We only want FORECAST sales for today and future
        date_filter_ventas += " AND CAST(fecha_proyeccion AS DATE) >= CAST(GETDATE() AS DATE)"

        # Prepare params in exact order of CTE execution
        params = p_pagos + p_ventas + p_ventas_real + p_gastos + p_gastos

        query = f"""
            WITH Pagos AS (
                SELECT 
                    CAST(COALESCE(PP.FechaPlanificada, SAACXP.FechaV) AS DATE) AS Fecha,
                    SUM(SAACXP.Saldo) AS SalidasBs,
                    SUM(SAACXP.Saldo / NULLIF(dt_emision.dolarbcv, 0)) AS SalidasPagosUSD
                FROM dbo.SAACXP WITH (NOLOCK)
                LEFT JOIN EnterpriseAdmin_AMC.Procurement.PagosPlanificados PP WITH (NOLOCK) ON SAACXP.NroUnico = PP.NroUnico
                OUTER APPLY (
                    SELECT TOP 1 dolarbcv 
                    FROM EnterpriseAdmin_AMC.dbo.dolartoday WITH (NOLOCK)
                    WHERE CAST(fecha AS DATE) <= CAST(SAACXP.FechaE AS DATE)
                    ORDER BY fecha DESC
                ) dt_emision
                WHERE SAACXP.Saldo > 0 AND {date_filter_pagos}
                GROUP BY CAST(COALESCE(PP.FechaPlanificada, SAACXP.FechaV) AS DATE)
            ),
            Ventas AS (
                SELECT 
                    CAST(DATEADD(day, {delay_days}, fecha_proyeccion) AS DATE) AS Fecha,
                    SUM(monto_proyectado_ves) AS EntradasBs,
                    SUM(monto_proyectado_usd) AS EntradasUSD
                FROM EnterpriseAdmin_AMC.Procurement.sales_forecast WITH (NOLOCK)
                WHERE {date_filter_ventas}
                GROUP BY CAST(DATEADD(day, {delay_days}, fecha_proyeccion) AS DATE)
                UNION ALL
                SELECT
                    CAST(DATEADD(day, {delay_days}, fecha) AS DATE) AS Fecha,
                    SUM(MtoVentas) AS EntradasBs,
                    SUM(CAST(MtoDolar AS FLOAT)) AS EntradasUSD
                FROM EnterpriseAdmin_AMC.dbo.CUSTOM_SAEVTA WITH (NOLOCK)
                WHERE {date_filter_ventas_real}
                GROUP BY CAST(DATEADD(day, {delay_days}, fecha) AS DATE)
            ),
            GastosFarmacia AS (
                SELECT 
                    CAST(fecha_proyectada AS DATE) AS Fecha,
                    SUM(monto_usd) AS SalidasFarmaciaUSD
                FROM EnterpriseAdmin_AMC.Procurement.GastosProgramados WITH (NOLOCK)
                WHERE tipo = 'Farmacia' AND estado = 'Pendiente' AND {date_filter_gastos}
                GROUP BY CAST(fecha_proyectada AS DATE)
            ),
            GastosPersonales AS (
                SELECT 
                    CAST(fecha_proyectada AS DATE) AS Fecha,
                    SUM(monto_usd) AS SalidasPersonalesUSD
                FROM EnterpriseAdmin_AMC.Procurement.GastosProgramados WITH (NOLOCK)
                WHERE tipo = 'Personal' AND estado = 'Pendiente' AND {date_filter_gastos}
                GROUP BY CAST(fecha_proyectada AS DATE)
            ),
            Fechas AS (
                SELECT Fecha FROM Pagos
                UNION
                SELECT Fecha FROM Ventas
                UNION
                SELECT Fecha FROM GastosFarmacia
                UNION
                SELECT Fecha FROM GastosPersonales
            ),
            ResumenDiario AS (
                SELECT 
                    FORMAT(F.Fecha, 'yyyy-MM-dd') AS Periodo,
                    F.Fecha AS RealFecha,
                    CAST(
                        CASE 
                            WHEN FORMAT(F.Fecha, 'yyyy-MM-dd') = '{fecha_arranque}' THEN 0 
                            ELSE ISNULL(ROUND(SUM(V.EntradasUSD), 2), 0) 
                        END AS FLOAT
                    ) AS EntradasUSD,
                    CAST(ISNULL(ROUND(SUM(P.SalidasPagosUSD), 2), 0) AS FLOAT) AS SalidasPagosUSD,
                    CAST(ISNULL(ROUND(SUM(GF.SalidasFarmaciaUSD), 2), 0) AS FLOAT) AS SalidasFarmaciaUSD,
                    CAST(ISNULL(ROUND(SUM(GP.SalidasPersonalesUSD), 2), 0) AS FLOAT) AS SalidasPersonalesUSD
                FROM Fechas F
                LEFT JOIN Ventas V ON F.Fecha = V.Fecha
                LEFT JOIN Pagos P ON F.Fecha = P.Fecha
                LEFT JOIN GastosFarmacia GF ON F.Fecha = GF.Fecha
                LEFT JOIN GastosPersonales GP ON F.Fecha = GP.Fecha
                GROUP BY F.Fecha
            )
            SELECT 
                Periodo,
                EntradasUSD,
                SalidasPagosUSD,
                SalidasFarmaciaUSD,
                SalidasPersonalesUSD,
                
                -- Saldo Real Acumulado usando SUM() OVER()
                CAST(ROUND(
                    -- Caja Inicial Total en USD (Caja USD + CajaBs convertido)
                    ( ? + ( ? / NULLIF((SELECT TOP 1 dolarbcv FROM EnterpriseAdmin_AMC.dbo.dolartoday WHERE CAST(fecha AS DATE) <= CAST(? AS DATE) ORDER BY fecha DESC), 0) ) )
                    + SUM(EntradasUSD - SalidasPagosUSD - SalidasFarmaciaUSD - SalidasPersonalesUSD) OVER (ORDER BY RealFecha ROWS UNBOUNDED PRECEDING)
                , 2) AS FLOAT) AS SaldoRealCajaUSD
                
            FROM ResumenDiario
            ORDER BY RealFecha
        """
        
        # Add the parameters for the Caja calculation at the end: Caja USD, Caja Bs, Fecha Arranque Tasa
        params.extend([caja_usd, caja_bs, fecha_arranque])

        cursor.execute(query, params)
        columns = [column[0] for column in cursor.description]
        results = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # Apply strict UI date filtering since the backend calculated from Day Zero to get correct rolling totals
        filtered_results = []
        if not desde:
            desde = '1900-01-01'
            
        for row in results:
            if row['Periodo'] >= desde:
                if hasta:
                    if row['Periodo'] <= hasta:
                        filtered_results.append(row)
                else:
                    filtered_results.append(row)
            
        return {"data": filtered_results}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@app.get("/api/forecast-events")
async def get_forecast_events():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, FORMAT(fecha, 'yyyy-MM-dd') as fecha, tipo_evento, valor FROM EnterpriseAdmin_AMC.Procurement.forecast_events ORDER BY fecha DESC")
        columns = [column[0] for column in cursor.description]

        return {"data": [dict(zip(columns, row)) for row in cursor.fetchall()]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@app.post("/api/forecast-events")
async def add_forecast_event(event: ForecastEventRequest):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        query = "INSERT INTO EnterpriseAdmin_AMC.Procurement.forecast_events (fecha, tipo_evento, valor) VALUES (?, ?, ?)"
        cursor.execute(query, (event.fecha, event.tipo_evento, event.valor))
        conn.commit()

        return {"message": "Event added successfully"}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

# --- Rutas de Plantillas de Gastos ---

@app.get("/api/expense-templates")
async def get_expense_templates():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, descripcion, tipo, monto_estimado_usd as monto_usd, dia_mes_estimado FROM EnterpriseAdmin_AMC.Procurement.PlantillasGastos ORDER BY tipo, descripcion")
        columns = [column[0] for column in cursor.description]
        return {"data": [dict(zip(columns, row)) for row in cursor.fetchall()]}
    except Exception as e:
        logging.error(f"Error in get_expense_templates: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@app.post("/api/expense-templates")
async def save_expense_template(template: ExpenseTemplateRequest):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        if template.id:
            query = """
                UPDATE EnterpriseAdmin_AMC.Procurement.PlantillasGastos 
                SET descripcion=?, tipo=?, monto_estimado_usd=?, dia_mes_estimado=? 
                WHERE id=?
            """
            cursor.execute(query, (template.descripcion, template.tipo, template.monto_estimado_usd, template.dia_mes_estimado, template.id))
        else:
            query = """
                INSERT INTO EnterpriseAdmin_AMC.Procurement.PlantillasGastos 
                (descripcion, tipo, monto_estimado_usd, dia_mes_estimado) 
                VALUES (?, ?, ?, ?)
            """
            cursor.execute(query, (template.descripcion, template.tipo, template.monto_estimado_usd, template.dia_mes_estimado))
            
        conn.commit()
        return {"message": "Plantilla guardada exitosamente"}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@app.delete("/api/expense-templates/{id}")
async def delete_expense_template(id: int):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM EnterpriseAdmin_AMC.Procurement.PlantillasGastos WHERE id = ?", (id,))
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Plantilla no encontrada")
        conn.commit()
        return {"message": "Plantilla eliminada"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

# --- Rutas de Modulo Batch/Generación Mensual ---

@app.get("/api/expenses/generate-batch/{mes}/{anio}")
async def get_expense_batch(mes: int, anio: int):
    """
    Simula qué debería pagarse en base a las plantillas y el mes para que la UI lo verifique.
    """
    try:
        from datetime import date
        pass  # Just ensuring it works locally
    except:
        pass
    
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, descripcion, tipo, monto_estimado_usd, dia_mes_estimado FROM EnterpriseAdmin_AMC.Procurement.PlantillasGastos")
        
        batch = []
        for row in cursor.fetchall():
            dia = row.dia_mes_estimado
            # Limitar dia al máximo de días del mes si es un día como 31 y el mes es febrero
            try:
                import calendar
                max_dia = calendar.monthrange(anio, mes)[1]
                dia = min(dia, max_dia)
                fecha_proy = f"{anio}-{mes:02d}-{dia:02d}"
            except:
                fecha_proy = f"{anio}-{mes:02d}-15" # Fallback
                
            batch.append({
                "template_id": row.id,
                "descripcion": row.descripcion,
                "tipo": row.tipo,
                "monto_usd": row.monto_estimado_usd,
                "fecha_proyectada": fecha_proy,
                "estado": "Pendiente"
            })
            
        return {"data": batch}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@app.post("/api/expenses/batch")
async def save_expense_batch(payload: BatchExpenseRequest):
    """
    Inserta o limpia el mes y guarda los gastos confirmados desde la interfaz.
    """
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Elimina lo que sea de ese mes para evitar duplicados si el usuario regenera y guarda.
        # But only delete the generated templates (is_adhoc = 0 or NULL), not the individual varianbles.
        cursor.execute("DELETE FROM EnterpriseAdmin_AMC.Procurement.GastosProgramados WHERE YEAR(fecha_proyectada) = ? AND MONTH(fecha_proyectada) = ? AND (is_adhoc = 0 OR is_adhoc IS NULL)", (payload.anio, payload.mes))
        
        # Insertar los nuevos generados de la UI
        for e in payload.gastos:
            cursor.execute(
                "INSERT INTO EnterpriseAdmin_AMC.Procurement.GastosProgramados (descripcion, tipo, monto_usd, fecha_proyectada, estado, is_adhoc) VALUES (?, ?, ?, ?, ?, 0)",
                (e.descripcion, e.tipo, e.monto_usd, e.fecha_proyectada, e.estado)
            )
            
        conn.commit()
        return {"message": "Lote mensual guardado exitosamente"}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@app.post("/api/expenses/programmed/single")
async def save_single_expense(expense: ProgrammedExpense):
    """
    Guarda un gasto variable on-the-fly directamente a la BD y lo marca como adhoc.
    """
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO EnterpriseAdmin_AMC.Procurement.GastosProgramados (descripcion, tipo, monto_usd, fecha_proyectada, estado, is_adhoc) VALUES (?, ?, ?, ?, ?, 1)",
            (expense.descripcion, expense.tipo, expense.monto_usd, expense.fecha_proyectada, expense.estado)
        )
        conn.commit()
        return {"message": "Gasto ad-hoc guardado exitosamente"}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()
        
@app.get("/api/expenses/programmed")
async def get_programmed_expenses(mes: int = None, anio: int = None):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        where_clause = ""
        params = []
        if mes and anio:
             where_clause = "WHERE YEAR(fecha_proyectada) = ? AND MONTH(fecha_proyectada) = ?"
             params.extend([anio, mes])
        
        query = f"SELECT id, descripcion, tipo, monto_usd, FORMAT(fecha_proyectada, 'yyyy-MM-dd') as fecha_proyectada, estado FROM EnterpriseAdmin_AMC.Procurement.GastosProgramados {where_clause} ORDER BY fecha_proyectada ASC"
        cursor.execute(query, tuple(params))
        columns = [column[0] for column in cursor.description]
        return {"data": [dict(zip(columns, row)) for row in cursor.fetchall()]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()
        
@app.delete("/api/expenses/programmed/{id}")
async def delete_programmed_expense(id: int):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM EnterpriseAdmin_AMC.Procurement.GastosProgramados WHERE id = ?", (id,))
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Gasto no encontrado")
        conn.commit()
        return {"message": "Gasto eliminado"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@app.delete("/api/forecast-events/{event_id}")
async def delete_forecast_event(event_id: int):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        query = "DELETE FROM EnterpriseAdmin_AMC.Procurement.forecast_events WHERE id = ?"
        cursor.execute(query, (event_id,))

        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Event not found")
        conn.commit()
        return {"message": "Event deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@app.get("/api/export/{report_type}")
async def export_xlsx(report_type: str, desde: Optional[str] = None, hasta: Optional[str] = None):
    try:
        data = []
        filename = f"{report_type}_reporte.xlsx"

        if report_type == "cuentas-por-pagar":
            res = await get_cuentas_por_pagar("", desde, hasta)
            data = res.get("data", [])
        elif report_type == "aging":
            res = await report_aging()
            data = res.get("data", [])
        elif report_type == "compras":
            res = await report_compras(desde, hasta)
            data = res.get("data", [])
        elif report_type == "debit-notes":
            res = await get_debit_notes()
            data = res.get("data", [])
        else:
            raise HTTPException(status_code=400, detail="Invalid report type")

        if not data:
            raise HTTPException(status_code=404, detail="No data to export")

        import openpyxl  # type: ignore
        from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type[:31]

        # Header row with styling
        headers = list(data[0].keys())
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, key in enumerate(headers, start=1):
                val = row_data.get(key)
                # Convert date/datetime to string for readability
                if hasattr(val, 'isoformat'):
                    val = str(val)[0:19]  # type: ignore
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Auto-width columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(int(max_len) + 3, 40)

        # Save to bytes buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error exporting XLSX: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ==========================================
# RETENCIONES DE IVA MODULE
# ==========================================

@app.get("/api/retenciones/config")
async def get_retenciones_config():
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT RIF_Agente, RazonSocial_Agente, DireccionFiscal_Agente, ValorUT, UltimoSecuencial FROM EnterpriseAdmin_AMC.Procurement.Retenciones_Config WHERE Id = 1")
        row = cursor.fetchone()
        if not row:
            return {"data": {}}
        
        data = {
            "RifAgente": row[0],
            "NombreAgente": row[1],
            "DireccionAgente": row[2],
            "ValorUT": float(row[3]) if row[3] else 0,
            "ProximoSecuencial": row[4]
        }
        return {"data": data}
    except Exception as e:
        logging.error(f"Error checking config: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@app.put("/api/retenciones/config")
async def update_retenciones_config(payload: dict = Body(...)):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE EnterpriseAdmin_AMC.Procurement.Retenciones_Config 
            SET RIF_Agente = ?, RazonSocial_Agente = ?, DireccionFiscal_Agente = ?, ValorUT = ?, UltimoSecuencial = ?
            WHERE Id = 1
        """, (
            payload.get("RifAgente"), 
            payload.get("NombreAgente"), 
            payload.get("DireccionAgente"), 
            float(payload.get("ValorUT", 0)),
            int(payload.get("ProximoSecuencial", 0))
        ))
        conn.commit()
        return {"message": "Configuración actualizada"}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        logging.error(f"Error updating retenciones config: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@app.get("/api/retenciones")
async def get_retenciones(desde: Optional[str] = None, hasta: Optional[str] = None):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        query = "SELECT * FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA WHERE 1=1"
        params = []
        if desde:
            query += " AND FechaRetencion >= ?"
            params.append(desde + " 00:00:00")
        if hasta:
            query += " AND FechaRetencion <= ?"
            params.append(hasta + " 23:59:59")
            
        query += " ORDER BY FechaRetencion DESC, Id DESC"
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        columns = [column[0] for column in cursor.description]
        data = [dict(zip(columns, row)) for row in rows]
        return {"data": data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@app.post("/api/retenciones")
async def crear_retencion(payload: dict = Body(...)):
    """Create retention records. Supports batch: {FechaRetencion, facturas: [{NumeroD, CodProv, ...}, ...]}"""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # 1. Generate sequential number
        cursor.execute("SELECT UltimoSecuencial FROM EnterpriseAdmin_AMC.Procurement.Retenciones_Config WHERE Id = 1")
        last_seq = cursor.fetchone()[0]
        new_seq = last_seq + 1
        
        from datetime import datetime
        now = datetime.now()
        nro_comprobante = f"{now.strftime('%Y%m')}{str(new_seq).zfill(8)}"
        
        # Support batch (new format) or single (legacy format)
        facturas = payload.get("facturas", [payload])  # If no facturas key, treat payload as single invoice
        fecha_retencion = payload.get("FechaRetencion", now.strftime('%Y-%m-%d'))
        
        inserted_ids = []
        for f in facturas:
            cursor.execute("""
                INSERT INTO EnterpriseAdmin_AMC.Procurement.Retenciones_IVA 
                (NumeroComprobante, NumeroD, CodProv, FechaFactura, FechaRetencion, NroControl, 
                 MontoTotal, BaseImponible, MontoExento, Alicuota, IVACausado, PorcentajeRetencion, 
                 MontoRetenido, Estado, TipoOperacion, TipoDocumento, DocAfectado)
                OUTPUT INSERTED.Id
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'EMITIDO', '01', '01', NULL)
            """, (
                nro_comprobante, f["NumeroD"], f["CodProv"], 
                f.get("FechaFactura", fecha_retencion),
                fecha_retencion, f.get("NroControl", "00-000000"),
                float(f.get("MontoTotal", 0)), float(f.get("BaseImponible", 0)), 
                float(f.get("MontoExento", 0)),
                float(f.get("Alicuota", 16)), float(f.get("IVACausado", 0)), 
                float(f.get("PorcentajeRetencion", 75)),
                float(f.get("MontoRetenido", 0))
            ))
            new_id = cursor.fetchone()[0]
            inserted_ids.append(new_id)
        
        # 3. Update sequential
        cursor.execute("UPDATE EnterpriseAdmin_AMC.Procurement.Retenciones_Config SET UltimoSecuencial = ? WHERE Id = 1", (new_seq,))
        
        conn.commit()
        logging.info(f"Retención {nro_comprobante} creada con {len(facturas)} factura(s)")
        return {"message": "Retención creada exitosamente", "NumeroComprobante": nro_comprobante, "Ids": inserted_ids}
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        logging.error(f"Error creating retencion: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

@app.patch("/api/retenciones/{id_ret}")
async def anular_retencion(id_ret: int):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT Estado FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA WHERE Id = ?", (id_ret,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Retención no encontrada")
        if row[0] == 'ENTERADO':
            raise HTTPException(status_code=400, detail="No se puede anular una retención ENTERADA (declarada ante SENIAT)")
            
        cursor.execute("UPDATE EnterpriseAdmin_AMC.Procurement.Retenciones_IVA SET Estado = 'ANULADO' WHERE Id = ?", (id_ret,))
        conn.commit()
        return {"message": "Retención anulada"}
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

# --- PDF Generation for Retenciones ---
def generar_pdf_retencion(config: dict, retenciones: list) -> bytes:
    """Generate a PDF comprobante de retención IVA. Supports multiple invoices in one comprobante."""
    from fpdf import FPDF
    from datetime import datetime

    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    
    # Colors
    header_bg = (0, 51, 102)   # Dark blue
    header_fg = (255, 255, 255) # White
    row_alt = (230, 240, 250)  # Light blue
    
    # --- Header ---
    pdf.set_font('Helvetica', 'B', 14)
    pdf.set_fill_color(*header_bg)
    pdf.set_text_color(*header_fg)
    pdf.cell(0, 10, 'COMPROBANTE DE RETENCIÓN DEL IVA', 0, 1, 'C', fill=True)
    pdf.ln(3)
    
    # --- Agente info ---
    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_text_color(0, 0, 0)

    nro_comprobante = retenciones[0]["NumeroComprobante"] if retenciones else "N/A"
    fecha_ret = str(retenciones[0].get("FechaRetencion", ""))[:10] if retenciones else "N/A"

    info_labels = [
        ("Agente de Retención:", config.get("RazonSocial_Agente", "")),
        ("RIF:", config.get("RIF_Agente", "")),
        ("Dirección Fiscal:", config.get("DireccionFiscal_Agente", "")),
        ("Nro. Comprobante:", nro_comprobante),
        ("Fecha de Retención:", fecha_ret),
    ]
    for label, value in info_labels:
        pdf.set_font('Helvetica', 'B', 9)
        pdf.cell(45, 6, label, 0, 0)
        pdf.set_font('Helvetica', '', 9)
        pdf.cell(0, 6, str(value), 0, 1)
    
    pdf.ln(3)
    
    # --- Sujeto Retenido (from first record) ---
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(45, 6, "Sujeto Retenido:", 0, 0)
    pdf.set_font('Helvetica', '', 9)
    prov_name = retenciones[0].get("ProveedorNombre") or retenciones[0].get("CodProv", "")
    pdf.cell(0, 6, str(prov_name), 0, 1)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(45, 6, "RIF Proveedor:", 0, 0)
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 6, str(retenciones[0].get("CodProv", "")), 0, 1)
    pdf.ln(5)
    
    # --- Table Header ---
    cols = [
        ("Factura", 25), ("Nro Control", 25), ("Fecha Factura", 22),
        ("Monto Total", 25), ("Base Imponible", 25), ("IVA %", 12),
        ("IVA Causado", 22), ("Ret. %", 12), ("Monto Retenido", 25)
    ]
    
    pdf.set_font('Helvetica', 'B', 7)
    pdf.set_fill_color(*header_bg)
    pdf.set_text_color(*header_fg)
    for name, width in cols:
        pdf.cell(width, 7, name, 1, 0, 'C', fill=True)
    pdf.ln()
    
    # --- Table Rows ---
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', '', 7)
    
    total_monto = 0
    total_base = 0
    total_iva = 0
    total_retenido = 0
    
    for i, r in enumerate(retenciones):
        if i % 2 == 1:
            pdf.set_fill_color(*row_alt)
            fill = True
        else:
            fill = False
            
        monto = float(r.get("MontoTotal", 0))
        base = float(r.get("BaseImponible", 0))
        alicuota = float(r.get("Alicuota", 0))
        iva = float(r.get("IVACausado", base * alicuota / 100))
        pct_ret = float(r.get("PorcentajeRetencion", 0))
        retenido = float(r.get("MontoRetenido", 0))
        
        total_monto += monto
        total_base += base
        total_iva += iva
        total_retenido += retenido
        
        fecha_fact = str(r.get("FechaFactura", ""))[:10]
        
        row_data = [
            str(r.get("NumeroD", "")),
            str(r.get("NroControl", "")),
            fecha_fact,
            f"{monto:,.2f}",
            f"{base:,.2f}",
            f"{alicuota:.0f}%",
            f"{iva:,.2f}",
            f"{pct_ret:.0f}%",
            f"{retenido:,.2f}"
        ]
        for j, (name, width) in enumerate(cols):
            align = 'R' if j >= 3 else 'L'
            pdf.cell(width, 6, row_data[j], 1, 0, align, fill=fill)
        pdf.ln()
    
    # --- Totals ---
    pdf.set_font('Helvetica', 'B', 7)
    pdf.set_fill_color(220, 220, 220)
    
    # Empty cells for first 3 columns
    total_width = cols[0][1] + cols[1][1] + cols[2][1]
    pdf.cell(total_width, 7, 'TOTALES:', 1, 0, 'R', fill=True)
    pdf.cell(cols[3][1], 7, f"{total_monto:,.2f}", 1, 0, 'R', fill=True)
    pdf.cell(cols[4][1], 7, f"{total_base:,.2f}", 1, 0, 'R', fill=True)
    pdf.cell(cols[5][1], 7, "", 1, 0, 'C', fill=True)
    pdf.cell(cols[6][1], 7, f"{total_iva:,.2f}", 1, 0, 'R', fill=True)
    pdf.cell(cols[7][1], 7, "", 1, 0, 'C', fill=True)
    pdf.cell(cols[8][1], 7, f"{total_retenido:,.2f}", 1, 0, 'R', fill=True)
    pdf.ln(12)
    
    # --- Footer ---
    pdf.set_font('Helvetica', '', 8)
    pdf.cell(0, 5, f"Documento generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", 0, 1, 'C')
    
    return pdf.output()


@app.get("/api/retenciones/{id_ret}/pdf")
async def get_retencion_pdf(id_ret: int):
    """Generate and return PDF preview of a retention comprobante."""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Get config
        cursor.execute("SELECT RIF_Agente, RazonSocial_Agente, DireccionFiscal_Agente, ValorUT, UltimoSecuencial FROM EnterpriseAdmin_AMC.Procurement.Retenciones_Config WHERE Id = 1")
        cfg_row = cursor.fetchone()
        config = dict(zip([c[0] for c in cursor.description], cfg_row)) if cfg_row else {}
        
        # Get retention(s) - may be grouped by NumeroComprobante
        cursor.execute("""
            SELECT r.*, p.Descrip as ProveedorNombre
            FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA r
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON r.CodProv = p.CodProv
            WHERE r.Id = ?
        """, (id_ret,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Retención no encontrada")
        
        columns = [c[0] for c in cursor.description]
        main_ret = dict(zip(columns, row))
        
        # Fetch all retentions with same NumeroComprobante (for grouped comprobantes)
        cursor.execute("""
            SELECT r.*, p.Descrip as ProveedorNombre
            FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA r
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON r.CodProv = p.CodProv
            WHERE r.NumeroComprobante = ?
            ORDER BY r.Id
        """, (main_ret["NumeroComprobante"],))
        all_rows = cursor.fetchall()
        retenciones = [dict(zip(columns, r)) for r in all_rows]
        
        pdf_bytes = generar_pdf_retencion(config, retenciones)
        
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f"inline; filename=Retencion_{main_ret['NumeroComprobante']}.pdf"}
        )
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error generating PDF: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()


@app.post("/api/retenciones/{id_ret}/send-email")
async def send_retencion_email(id_ret: int):
    """Send retention comprobante via email to the provider, with PDF + Excel attachments."""
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Get config
        cursor.execute("SELECT RIF_Agente, RazonSocial_Agente, DireccionFiscal_Agente, ValorUT FROM EnterpriseAdmin_AMC.Procurement.Retenciones_Config WHERE Id = 1")
        cfg_row = cursor.fetchone()
        config = dict(zip([c[0] for c in cursor.description], cfg_row)) if cfg_row else {}
        
        # Get retention
        cursor.execute("""
            SELECT r.*, p.Descrip as ProveedorNombre
            FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA r
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON r.CodProv = p.CodProv
            WHERE r.Id = ?
        """, (id_ret,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Retención no encontrada")
        
        columns = [c[0] for c in cursor.description]
        main_ret = dict(zip(columns, row))
        
        # Fetch all grouped retentions
        cursor.execute("""
            SELECT r.*, p.Descrip as ProveedorNombre
            FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA r
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON r.CodProv = p.CodProv
            WHERE r.NumeroComprobante = ?
            ORDER BY r.Id
        """, (main_ret["NumeroComprobante"],))
        all_rows = cursor.fetchall()
        retenciones = [dict(zip(columns, r)) for r in all_rows]
        
        # Get provider email
        cursor.execute("""
            SELECT c.Email, p.Descrip 
            FROM EnterpriseAdmin_AMC.Procurement.ProveedorCondiciones c 
            LEFT JOIN EnterpriseAdmin_AMC.dbo.SAPROV p ON c.CodProv = p.CodProv 
            WHERE c.CodProv = ?
        """, (main_ret["CodProv"],))
        prov = cursor.fetchone()
        
        if not prov or not prov.Email:
            return {"email_sent": False, "message": "Proveedor sin email configurado."}
        
        # Generate PDF
        pdf_bytes = generar_pdf_retencion(config, retenciones)
        
        # Generate styled Excel
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        excel_data = []
        for r in retenciones:
            excel_data.append({
                "Nro Factura": r["NumeroD"],
                "Nro Control": r.get("NroControl", ""),
                "Fecha Factura": str(r.get("FechaFactura", ""))[:10],
                "Monto Total (Bs)": float(r.get("MontoTotal", 0)),
                "Base Imponible (Bs)": float(r.get("BaseImponible", 0)),
                "Alícuota IVA (%)": float(r.get("Alicuota", 0)),
                "IVA Causado (Bs)": float(r.get("IVACausado", 0)),
                "% Retención": float(r.get("PorcentajeRetencion", 0)),
                "Monto Retenido (Bs)": float(r.get("MontoRetenido", 0))
            })
        
        df = pd.DataFrame(excel_data)
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Retención IVA')
            ws = writer.sheets['Retención IVA']
            
            # Header style
            header_font = Font(bold=True, color="FFFFFF", size=10)
            header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            # Data formatting
            for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row_cells:
                    cell.border = thin_border
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(horizontal="left")
            
            # Totals row
            total_row = ws.max_row + 1
            ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)
            ws.cell(row=total_row, column=1).border = thin_border
            ws.cell(row=total_row, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            for col_idx in range(2, ws.max_column + 1):
                cell = ws.cell(row=total_row, column=col_idx)
                cell.border = thin_border
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                # Sum numeric columns (4,5,7,9 are numeric)
                if col_idx in [4, 5, 7, 9]:
                    col_letter = chr(64 + col_idx)
                    cell.value = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
                    cell.number_format = '#,##0.00'
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="right")
            
            # Auto-fit column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 25)
        
        excel_buffer.seek(0)
        
        # Build email
        service = get_gmail_service()
        if not service:
            return {"email_sent": False, "message": "Gmail API no disponible."}
        
        # Support multiple emails separated by ;
        emails = [e.strip() for e in prov.Email.split(";") if e.strip()]
        remitente = os.getenv("SMTP_EMAIL", "")
        
        nro_comp = main_ret["NumeroComprobante"]
        prov_nombre = prov.Descrip or "Proveedor"
        
        msg = MIMEMultipart()
        msg['From'] = remitente
        msg['To'] = ", ".join(emails)
        msg['Subject'] = f"Comprobante de Retención IVA - {nro_comp}"
        
        total_retenido = sum(float(r.get("MontoRetenido", 0)) for r in retenciones)
        
        cuerpo = f"""Estimados/as {prov_nombre},

Adjunto a este correo el comprobante de retención de IVA Nro. {nro_comp}.
Monto total retenido: Bs. {total_retenido:,.2f}

Atentamente,
El equipo de Administración."""
        msg.attach(MIMEText(cuerpo, 'plain', 'utf-8'))
        
        # Attach PDF
        part_pdf = MIMEBase("application", "pdf")
        part_pdf.set_payload(pdf_bytes)
        encoders.encode_base64(part_pdf)
        part_pdf.add_header("Content-Disposition", f"attachment; filename=Retencion_{nro_comp}.pdf")
        msg.attach(part_pdf)
        
        # Attach Excel
        part_excel = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part_excel.set_payload(excel_buffer.read())
        encoders.encode_base64(part_excel)
        part_excel.add_header("Content-Disposition", f"attachment; filename=Resumen_Retencion_{nro_comp}.xlsx")
        msg.attach(part_excel)
        
        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
        service.users().messages().send(userId='me', body={'raw': raw}).execute()
        logging.info(f"Correo retención {nro_comp} enviado a {', '.join(emails)}")
        
        return {"email_sent": True, "message": f"Correo enviado a {', '.join(emails)}"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error sending retention email: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()


@app.get("/api/retenciones/export-txt")
async def export_retenciones_txt(desde: Optional[str] = None, hasta: Optional[str] = None):
    try:
        conn = database.get_db_connection()
        cursor = conn.cursor()
        
        # Get Agent RIF
        cursor.execute("SELECT RIF_Agente FROM EnterpriseAdmin_AMC.Procurement.Retenciones_Config WHERE Id = 1")
        rif_agente = cursor.fetchone()[0]
        
        query = "SELECT * FROM EnterpriseAdmin_AMC.Procurement.Retenciones_IVA WHERE Estado = 'EMITIDO'"
        params = []
        if desde:
            query += " AND FechaRetencion >= ?"
            params.append(desde + " 00:00:00")
        if hasta:
            query += " AND FechaRetencion <= ?"
            params.append(hasta + " 23:59:59")
            
        query += " ORDER BY FechaRetencion ASC"
        cursor.execute(query, params)
        rows = cursor.fetchall()
        columns = [column[0] for column in cursor.description]
        data = [dict(zip(columns, row)) for row in rows]
        
        if not data:
            raise HTTPException(status_code=404, detail="No hay retenciones para exportar en este período")
            
        lines = []
        for r in data:
            # Format: 16 columns tab separated
            # 1: RIF Agente
            # 2: Periodo (AAAAMM)
            # 3: Fecha Doc (AAAA-MM-DD)
            # 4: Tipo Operacion (01, 02, 03)
            # 5: Tipo Doc (01, 02, 03)
            # 6: RIF Prov
            # 7: Nro Factura
            # 8: Nro Control
            # 9: Monto Total
            # 10: Base Imponible
            # 11: Monto Retenido
            # 12: Doc Afectado
            # 13: Nro Comprobante
            # 14: Monto Exento
            # 15: Alicuota
            # 16: Nro Expediente
            
            periodo = str(r["FechaRetencion"])[0:7].replace("-", "") # type: ignore
            fdoc = str(r["FechaFactura"])[0:10] # type: ignore
            
            line = [
                rif_agente,
                periodo,
                fdoc,
                r["TipoOperacion"],
                r["TipoDocumento"],
                str(r["CodProv"]).replace("-", ""),  # clean RIF # type: ignore
                r["NumeroD"],
                r["NroControl"],
                f"{r['MontoTotal']:.2f}",
                f"{r['BaseImponible']:.2f}",
                f"{r['MontoRetenido']:.2f}",
                str(r["DocAfectado"] or "0"),
                str(r["NumeroComprobante"]),
                f"{r['MontoExento']:.2f}",
                f"{r['Alicuota']:.2f}",
                str(r["NroExpediente"] or "0")
            ]
            lines.append("\t".join([str(x) for x in line]))
            
            # Also mark them as ENTERADO
            cursor.execute("UPDATE EnterpriseAdmin_AMC.Procurement.Retenciones_IVA SET Estado = 'ENTERADO' WHERE Id = ?", (r["Id"],))
            
        conn.commit()
        
        txt_content = "\n".join(lines)
        return StreamingResponse(
            io.BytesIO(txt_content.encode("utf-8")),
            media_type="text/plain",
            headers={"Content-Disposition": f"attachment; filename=retenciones_{periodo}.txt"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        if 'conn' in locals(): conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals(): conn.close()

if __name__ == "__main__":
    import uvicorn # type: ignore
    uvicorn.run(app, host="0.0.0.0", port=8080)
